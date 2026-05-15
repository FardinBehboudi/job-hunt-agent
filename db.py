"""
db.py — SQLite data layer for the job hunt agent.

Database: uploads/jobhunt.db  (project-local, gitignored).
"""

import sqlite3
from contextlib import contextmanager
from datetime import date
from pathlib import Path

import pandas as pd

_DB_PATH = Path(__file__).resolve().parent / "uploads" / "jobhunt.db"


def _ensure_dir() -> None:
    _DB_PATH.parent.mkdir(exist_ok=True)


@contextmanager
def _conn():
    _ensure_dir()
    db = sqlite3.connect(str(_DB_PATH), timeout=15)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    try:
        yield db
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ── Schema ────────────────────────────────────────────────────────────────────

def init_db() -> None:
    with _conn() as db:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS applications (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                company          TEXT,
                role             TEXT,
                location         TEXT,
                date_applied     TEXT,
                status           TEXT,
                verdict          TEXT,
                match_pct        INTEGER,
                key_gap          TEXT,
                strengths        TEXT,
                company_size     TEXT,
                language         TEXT,
                job_url          TEXT UNIQUE,
                source           TEXT,
                interview_chance TEXT,
                archive_path     TEXT,
                created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS scraped_jobs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                title       TEXT,
                company     TEXT,
                location    TEXT,
                url         TEXT UNIQUE,
                description TEXT,
                source      TEXT,
                scraped_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS matched_jobs (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                scraped_job_id   INTEGER REFERENCES scraped_jobs(id),
                match_score      INTEGER,
                interview_chance TEXT,
                german_level     TEXT,
                skip_reason      TEXT,
                match_summary    TEXT
            );

            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS upcoming_interviews (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                date           TEXT,
                company        TEXT,
                role           TEXT,
                interview_type TEXT,
                time_berlin    TEXT,
                format         TEXT,
                job_url        TEXT,
                notes          TEXT,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS priority_tasks (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                priority   TEXT,
                company    TEXT,
                action     TEXT,
                deadline   TEXT,
                status     TEXT,
                notes      TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)


# ── Settings ──────────────────────────────────────────────────────────────────

def get_setting(key: str, default: str = "") -> str:
    with _conn() as db:
        row = db.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with _conn() as db:
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (key, value))


# ── Excel import ──────────────────────────────────────────────────────────────

# Column name aliases from the existing tracker spreadsheet
_IMPORT_ALIASES: dict[str, list[str]] = {
    "company":          ["company", "employer"],
    "role":             ["role", "job title", "title", "position"],
    "location":         ["location", "city"],
    "date_applied":     ["date applied", "date", "application date"],
    "status":           ["status", "application status"],
    "match_pct":        ["match score", "match %", "score"],
    "job_url":          ["job url", "url", "link"],
    "source":           ["source"],
    "interview_chance": ["interview chance", "chance"],
    "archive_path":     ["archive folder", "archive", "folder"],
    "language":         ["german required", "german level", "german"],
}


def _alias_val(row: dict, col_lower: dict[str, str], key: str):
    for alias in _IMPORT_ALIASES.get(key, [key]):
        col = col_lower.get(alias.lower())
        if col is not None:
            v = row.get(col)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            if hasattr(v, "strftime"):
                return v.strftime("%Y-%m-%d")
            s = str(v).strip()
            if s:
                return s
    return None


def _parse_match_pct(raw) -> "int | None":
    if raw is None:
        return None
    try:
        v = float(str(raw).replace("%", "").strip())
        return round(v * 100) if 0 < v <= 1.0 else round(v)
    except (ValueError, TypeError):
        return None


def import_from_excel(filepath: "str | Path") -> int:
    """
    Read Applied / Interviews / Rejected sheets (or a single sheet with a
    Status column) and INSERT rows into applications.  Returns count inserted.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    init_db()
    with pd.ExcelFile(filepath) as xl:
        target_sheets = [s for s in ["Applied", "Interviews", "Rejected", "Offer"]
                         if s in xl.sheet_names]

        if not target_sheets:
            # Single-sheet workbook — infer status from Status column
            df = xl.parse(xl.sheet_names[0])
            df = df.where(pd.notna(df), other=None)
            return _import_df(df, status_override=None)

        total = 0
        for sheet in target_sheets:
            df = xl.parse(sheet)
            df = df.where(pd.notna(df), other=None)
            total += _import_df(df, status_override=sheet)
    return total


def _import_df(df: pd.DataFrame, status_override: "str | None") -> int:
    df.columns = [str(c).strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}
    inserted = 0
    with _conn() as db:
        for _, row in df.iterrows():
            r = row.to_dict()
            company = _alias_val(r, col_lower, "company")
            role    = _alias_val(r, col_lower, "role")
            if not company and not role:
                continue  # empty row
            status = status_override or _alias_val(r, col_lower, "status") or "Applied"
            try:
                db.execute("""
                    INSERT OR IGNORE INTO applications
                    (company, role, location, date_applied, status, match_pct,
                     job_url, source, interview_chance, archive_path, language)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    company,
                    role,
                    _alias_val(r, col_lower, "location"),
                    _alias_val(r, col_lower, "date_applied"),
                    status,
                    _parse_match_pct(_alias_val(r, col_lower, "match_pct")),
                    _alias_val(r, col_lower, "job_url"),
                    _alias_val(r, col_lower, "source"),
                    _alias_val(r, col_lower, "interview_chance"),
                    _alias_val(r, col_lower, "archive_path"),
                    _alias_val(r, col_lower, "language"),
                ))
                inserted += db.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass
    return inserted


# ── Dashboard queries ─────────────────────────────────────────────────────────

def get_overview_data() -> dict:
    """Return stats + upcoming scheduled events + priority pending tasks."""
    with _conn() as db:
        status_rows = db.execute(
            "SELECT status, COUNT(*) AS cnt FROM applications GROUP BY status"
        ).fetchall()

        upcoming = db.execute("""
            SELECT id, company, role, date_applied, status, job_url
            FROM applications
            WHERE status IN (
                'Call Scheduled ✓',
                'Technical Scheduled ✓',
                'Final Scheduled ✓'
            )
            ORDER BY date_applied DESC
            LIMIT 10
        """).fetchall()

        tasks = db.execute("""
            SELECT id, company, role, date_applied, status, job_url
            FROM applications
            WHERE status IN (
                '⏸️ Technical — awaiting your confirmation',
                '⏸️ Final — awaiting your confirmation',
                'Applied — unconfirmed'
            )
            ORDER BY date_applied ASC
        """).fetchall()

    buckets = {"Applied": 0, "Interviews": 0, "Rejected": 0, "Offer": 0}
    total = 0
    for r in status_rows:
        total += r["cnt"]
        sl = (r["status"] or "").lower()
        if "rejected" in sl:
            buckets["Rejected"] += r["cnt"]
        elif "offer" in sl:
            buckets["Offer"] += r["cnt"]
        elif any(k in sl for k in ("scheduled", "interview", "technical", "final", "call")):
            buckets["Interviews"] += r["cnt"]
        else:
            buckets["Applied"] += r["cnt"]

    return {
        "stats":           {"total": total, **buckets},
        "upcoming_events": [dict(r) for r in upcoming],
        "priority_tasks":  [dict(r) for r in tasks],
    }


def get_dashboard_data() -> dict:
    with _conn() as db:
        rows = db.execute("""
            SELECT id, company, role, location, date_applied, status,
                   verdict, match_pct, key_gap, strengths, company_size,
                   language, job_url, source, interview_chance, archive_path
            FROM applications
            ORDER BY date_applied DESC, created_at DESC
        """).fetchall()
        import_done = db.execute(
            "SELECT value FROM settings WHERE key='import_done'"
        ).fetchone()

    all_rows = [dict(r) for r in rows]
    statuses = ["Applied", "Interviews", "Rejected", "Offer"]
    grouped: dict[str, list] = {s: [] for s in statuses}
    for r in all_rows:
        s = r.get("status") or "Applied"
        grouped.setdefault(s, []).append(r)

    overview = {"total": len(all_rows)}
    for s in statuses:
        overview[s] = len(grouped.get(s, []))

    return {
        "overview":    overview,
        "Applied":     grouped.get("Applied", []),
        "Interviews":  grouped.get("Interviews", []),
        "Rejected":    grouped.get("Rejected", []),
        "Offer":       grouped.get("Offer", []),
        "import_done": import_done is not None,
    }


# ── Applications ──────────────────────────────────────────────────────────────

def get_applied_urls() -> set[str]:
    with _conn() as db:
        rows = db.execute(
            "SELECT job_url FROM applications WHERE job_url IS NOT NULL"
        ).fetchall()
    return {r["job_url"] for r in rows}


def is_already_applied(url: str) -> bool:
    if not url:
        return False
    with _conn() as db:
        row = db.execute(
            "SELECT id FROM applications WHERE job_url=?", (url,)
        ).fetchone()
    return row is not None


def log_application(job: dict, archive_path: str = "", status: str = "Applied") -> None:
    url = (job.get("url") or "").strip() or None
    with _conn() as db:
        if url:
            db.execute("""
                INSERT INTO applications
                (company, role, location, date_applied, status, match_pct,
                 job_url, source, interview_chance, archive_path, language)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(job_url) DO UPDATE SET
                    status       = excluded.status,
                    archive_path = excluded.archive_path,
                    date_applied = excluded.date_applied
            """, (
                job.get("company", ""),
                job.get("title",   ""),
                job.get("location", ""),
                date.today().isoformat(),
                status,
                job.get("match_score"),
                url,
                job.get("source", ""),
                job.get("interview_chance", ""),
                str(archive_path) if archive_path else "",
                job.get("german_level_required", ""),
            ))
        else:
            db.execute("""
                INSERT INTO applications
                (company, role, location, date_applied, status, match_pct,
                 source, interview_chance, archive_path, language)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                job.get("company", ""),
                job.get("title",   ""),
                job.get("location", ""),
                date.today().isoformat(),
                status,
                job.get("match_score"),
                job.get("source", ""),
                job.get("interview_chance", ""),
                str(archive_path) if archive_path else "",
                job.get("german_level_required", ""),
            ))


# ── Scraped / matched jobs ────────────────────────────────────────────────────

def insert_scraped_jobs(jobs: list[dict]) -> int:
    inserted = 0
    with _conn() as db:
        for job in jobs:
            url = (job.get("url") or "").strip()
            if not url:
                continue
            try:
                db.execute("""
                    INSERT OR IGNORE INTO scraped_jobs
                    (title, company, location, url, description, source)
                    VALUES (?,?,?,?,?,?)
                """, (
                    job.get("title", ""),
                    job.get("company", ""),
                    job.get("location", ""),
                    url,
                    job.get("description", ""),
                    job.get("source", "LinkedIn"),
                ))
                inserted += db.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass
    return inserted


def get_unmatched_scraped_jobs() -> list[dict]:
    with _conn() as db:
        rows = db.execute("""
            SELECT s.id, s.title, s.company, s.location, s.url, s.description, s.source
            FROM scraped_jobs s
            LEFT JOIN matched_jobs m ON m.scraped_job_id = s.id
            WHERE m.id IS NULL
        """).fetchall()
    return [dict(r) for r in rows]


def insert_matched_job(scraped_job_id: int, scores: dict) -> None:
    with _conn() as db:
        db.execute("""
            INSERT INTO matched_jobs
            (scraped_job_id, match_score, interview_chance,
             german_level, skip_reason, match_summary)
            VALUES (?,?,?,?,?,?)
        """, (
            scraped_job_id,
            scores.get("match_score", 0),
            scores.get("interview_chance", "low"),
            scores.get("german_level_required", "none"),
            scores.get("skip_reason"),
            scores.get("match_summary", ""),
        ))


def get_matched_jobs_for_apply(cfg: dict) -> list[dict]:
    min_score   = cfg.get("min_match_score", 70)
    skip_levels = {lvl.lower() for lvl in cfg.get("skip_german_levels", [])}
    with _conn() as db:
        rows = db.execute("""
            SELECT s.id  AS scraped_id,
                   s.title, s.company, s.location, s.url, s.description, s.source,
                   m.match_score, m.interview_chance,
                   m.german_level, m.skip_reason, m.match_summary
            FROM matched_jobs m
            JOIN scraped_jobs s ON s.id = m.scraped_job_id
            WHERE m.match_score >= ?
              AND m.skip_reason IS NULL
              AND (s.url IS NULL OR s.url NOT IN
                   (SELECT job_url FROM applications WHERE job_url IS NOT NULL))
            ORDER BY m.match_score DESC
        """, (min_score,)).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        if (d.get("german_level") or "none").lower() in skip_levels:
            continue
        result.append({
            "title":                 d["title"],
            "company":               d["company"],
            "location":              d["location"],
            "url":                   d["url"],
            "description":           d["description"],
            "source":                d["source"],
            "match_score":           d["match_score"],
            "interview_chance":      d["interview_chance"],
            "german_level_required": d["german_level"],
            "match_summary":         d["match_summary"],
        })
    return result


# ── Excel export ──────────────────────────────────────────────────────────────

def generate_excel_report() -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    with _conn() as db:
        rows = db.execute("""
            SELECT company, role, location, date_applied, status, match_pct,
                   interview_chance, language, job_url, archive_path, source
            FROM applications ORDER BY date_applied DESC, created_at DESC
        """).fetchall()

    all_rows = [dict(r) for r in rows]
    HEADERS = ["Company", "Role", "Location", "Date Applied", "Status",
               "Match %", "Interview Chance", "German", "Job URL", "Archive", "Source"]
    COL_WIDTHS = [20, 28, 16, 13, 14, 9, 16, 10, 45, 45, 12]

    def to_row(r):
        pct = r["match_pct"]
        return [
            r["company"], r["role"], r["location"], r["date_applied"], r["status"],
            f"{pct}%" if pct is not None else "",
            r["interview_chance"], r["language"], r["job_url"], r["archive_path"], r["source"],
        ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    grouped = {s: [r for r in all_rows if r["status"] == s]
               for s in ["Applied", "Interviews", "Rejected", "Offer"]}

    for sheet_name, sheet_rows in [
        ("Overview",   all_rows),
        ("Applied",    grouped["Applied"]),
        ("Interviews", grouped["Interviews"]),
        ("Rejected",   grouped["Rejected"]),
        ("Offer",      grouped["Offer"]),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        for i, cell in enumerate(ws[1], 0):
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(i + 1)].width = COL_WIDTHS[i]
        for r in sheet_rows:
            ws.append(to_row(r))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Upcoming interviews (manual) ──────────────────────────────────────────────

def get_upcoming_interviews() -> list[dict]:
    with _conn() as db:
        rows = db.execute(
            "SELECT * FROM upcoming_interviews ORDER BY date ASC, created_at ASC"
        ).fetchall()
    return [dict(r) for r in rows]


def add_interview(data: dict) -> int:
    with _conn() as db:
        cur = db.execute("""
            INSERT INTO upcoming_interviews
            (date, company, role, interview_type, time_berlin, format, job_url, notes)
            VALUES (?,?,?,?,?,?,?,?)
        """, (
            data.get("date", ""),
            data.get("company", ""),
            data.get("role", ""),
            data.get("interview_type", ""),
            data.get("time_berlin", ""),
            data.get("format", ""),
            data.get("job_url", ""),
            data.get("notes", ""),
        ))
        return cur.lastrowid


def update_interview(interview_id: int, data: dict) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE upcoming_interviews
            SET date=?, company=?, role=?, interview_type=?,
                time_berlin=?, format=?, job_url=?, notes=?
            WHERE id=?
        """, (
            data.get("date", ""),
            data.get("company", ""),
            data.get("role", ""),
            data.get("interview_type", ""),
            data.get("time_berlin", ""),
            data.get("format", ""),
            data.get("job_url", ""),
            data.get("notes", ""),
            interview_id,
        ))


def delete_interview(interview_id: int) -> None:
    with _conn() as db:
        db.execute("DELETE FROM upcoming_interviews WHERE id=?", (interview_id,))


# ── Priority tasks (manual) ────────────────────────────────────────────────────

def get_priority_tasks_manual() -> list[dict]:
    with _conn() as db:
        rows = db.execute(
            "SELECT * FROM priority_tasks ORDER BY created_at ASC"
        ).fetchall()
    return [dict(r) for r in rows]


def add_task_manual(data: dict) -> int:
    with _conn() as db:
        cur = db.execute("""
            INSERT INTO priority_tasks
            (priority, company, action, deadline, status, notes)
            VALUES (?,?,?,?,?,?)
        """, (
            data.get("priority", "NORMAL"),
            data.get("company", ""),
            data.get("action", ""),
            data.get("deadline", ""),
            data.get("status", ""),
            data.get("notes", ""),
        ))
        return cur.lastrowid


def update_task_manual(task_id: int, data: dict) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE priority_tasks
            SET priority=?, company=?, action=?, deadline=?, status=?, notes=?
            WHERE id=?
        """, (
            data.get("priority", "NORMAL"),
            data.get("company", ""),
            data.get("action", ""),
            data.get("deadline", ""),
            data.get("status", ""),
            data.get("notes", ""),
            task_id,
        ))


def delete_task_manual(task_id: int) -> None:
    with _conn() as db:
        db.execute("DELETE FROM priority_tasks WHERE id=?", (task_id,))
