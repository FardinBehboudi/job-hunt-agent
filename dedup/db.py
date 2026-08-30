"""
db.py — SQLite data layer for the job hunt agent.

Database: uploads/jobhunt.db  (project-local, gitignored).
"""

import shutil
import sqlite3
from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path

import pandas as pd

_DB_PATH = Path(__file__).resolve().parent.parent / "uploads" / "jobhunt.db"
_BACKUP_DIR = _DB_PATH.parent / "backups"
_KEEP_BACKUPS = 7  # daily copies to retain


def backup_db() -> Path:
    """Copy the live DB to uploads/backups/jobhunt_YYYYMMDD.db using SQLite's
    online-backup API so the copy is always consistent even under WAL mode.
    Returns the backup path. Keeps at most _KEEP_BACKUPS files (oldest deleted)."""
    _BACKUP_DIR.mkdir(exist_ok=True)
    dest = _BACKUP_DIR / f"jobhunt_{date.today().strftime('%Y%m%d')}.db"
    src = sqlite3.connect(str(_DB_PATH))
    dst = sqlite3.connect(str(dest))
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()
    # Prune old backups
    backups = sorted(_BACKUP_DIR.glob("jobhunt_*.db"))
    for old in backups[:-_KEEP_BACKUPS]:
        old.unlink(missing_ok=True)
    return dest


def _ensure_dir() -> None:
    _DB_PATH.parent.mkdir(exist_ok=True)


@contextmanager
def _conn():
    _ensure_dir()
    db = sqlite3.connect(str(_DB_PATH), timeout=15)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    try:
        yield db
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ── Schema ────────────────────────────────────────────────────────────────────

def init_db() -> None:
    with _conn() as db:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS applications (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                company          TEXT,
                role             TEXT,
                location         TEXT,
                date_applied     TEXT,
                status           TEXT,
                verdict          TEXT,
                match_pct        INTEGER,
                key_gap          TEXT,
                strengths        TEXT,
                company_size     TEXT,
                language         TEXT,
                job_url          TEXT UNIQUE,
                source           TEXT,
                interview_chance TEXT,
                archive_path     TEXT,
                created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS scraped_jobs (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                title        TEXT,
                company      TEXT,
                location     TEXT,
                url          TEXT UNIQUE,
                description  TEXT,
                source       TEXT,
                posted_date  TEXT DEFAULT '',
                has_easy_apply INTEGER DEFAULT 0,
                cache_status TEXT DEFAULT 'new',
                scraped_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS matched_jobs (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                scraped_job_id   INTEGER REFERENCES scraped_jobs(id),
                match_score      INTEGER,
                interview_chance TEXT,
                german_level     TEXT,
                skip_reason      TEXT,
                match_summary    TEXT
            );

            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS upcoming_interviews (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                date           TEXT,
                company        TEXT,
                role           TEXT,
                interview_type TEXT,
                time_berlin    TEXT,
                format         TEXT,
                job_url        TEXT,
                notes          TEXT,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS priority_tasks (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                priority   TEXT,
                company    TEXT,
                action     TEXT,
                deadline   TEXT,
                status     TEXT,
                notes      TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS seen_jobs (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                url                   TEXT UNIQUE,
                title                 TEXT,
                company               TEXT,
                location              TEXT,
                description           TEXT,
                source                TEXT,
                posted_date           TEXT DEFAULT '',
                first_scraped_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_scraped_at       TIMESTAMP,
                match_score           INTEGER DEFAULT NULL,
                interview_chance      TEXT DEFAULT NULL,
                skip_reason           TEXT DEFAULT NULL,
                german_level_required TEXT DEFAULT NULL,
                resume_hash           TEXT DEFAULT NULL,
                applied               INTEGER DEFAULT 0,
                dismissed             INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS manual_apply_queue (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                job_url         TEXT,
                title           TEXT,
                company         TEXT,
                platform        TEXT,
                note            TEXT,
                screenshot_path TEXT,
                created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status          TEXT DEFAULT 'pending'
            );

            CREATE TABLE IF NOT EXISTS apply_sessions (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                started_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                finished_at   TIMESTAMP,
                total_jobs    INTEGER DEFAULT 0,
                success_count INTEGER DEFAULT 0,
                manual_count  INTEGER DEFAULT 0,
                failed_count  INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS excluded_jobs (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                url         TEXT UNIQUE,
                company     TEXT,
                title       TEXT,
                reason      TEXT,
                excluded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS email_staging (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                email_uid             TEXT,
                email_message_id      TEXT UNIQUE,
                sender                TEXT,
                subject               TEXT,
                body_preview          TEXT,
                received_date         TEXT,
                source_folder         TEXT,
                matched_app_id        INTEGER REFERENCES applications(id),
                match_confidence      INTEGER DEFAULT 0,
                match_type            TEXT DEFAULT 'unmatched',
                predicted_folder      TEXT,
                confidence_score      INTEGER DEFAULT 0,
                classification_reason TEXT,
                status                TEXT DEFAULT 'pending',
                user_override_folder  TEXT,
                created_at            TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                reviewed_at           TIMESTAMP,
                executed_at           TIMESTAMP,
                notes                 TEXT
            );

            CREATE TABLE IF NOT EXISTS upcoming_events (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                app_id          INTEGER REFERENCES applications(id),
                event_type      TEXT,
                title           TEXT,
                description     TEXT,
                event_date      TEXT,
                event_time      TEXT,
                timezone        TEXT,
                priority        TEXT DEFAULT 'medium',
                source_email_id INTEGER REFERENCES email_staging(id),
                status          TEXT DEFAULT 'scheduled',
                created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS email_move_history (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                email_staging_id INTEGER REFERENCES email_staging(id),
                from_folder      TEXT,
                to_folder        TEXT,
                moved_at         TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                success          INTEGER DEFAULT 0,
                error_message    TEXT
            );
        """)
    # Migrate existing tables — add columns if missing
    _migrations = [
        ("scraped_jobs", "cache_status",          "TEXT DEFAULT 'new'"),
        ("scraped_jobs", "posted_date",           "TEXT DEFAULT ''"),
        ("scraped_jobs", "has_easy_apply",        "INTEGER DEFAULT 0"),
        ("seen_jobs",    "posted_date",           "TEXT DEFAULT ''"),
        ("seen_jobs",    "first_seen_at",         "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
        ("seen_jobs",    "first_scraped_at",      "TIMESTAMP"),
        ("seen_jobs",    "last_scraped_at",       "TIMESTAMP"),
        ("seen_jobs",    "german_level_required", "TEXT DEFAULT NULL"),
        ("seen_jobs",    "resume_hash",           "TEXT DEFAULT NULL"),
        ("seen_jobs",    "has_easy_apply",        "INTEGER DEFAULT 0"),
        # legacy — kept for existing rows; ignored in new code
        ("seen_jobs",             "match_summary",  "TEXT DEFAULT NULL"),
        ("manual_apply_queue",    "screenshot_path","TEXT DEFAULT NULL"),
        ("manual_apply_queue",    "session_id",     "TEXT DEFAULT NULL"),
        ("applications", "applied_by",  "TEXT DEFAULT ''"),
        ("applications", "apply_type", "TEXT DEFAULT ''"),
        ("applications", "last_email_date",       "TEXT DEFAULT NULL"),
        ("applications", "last_email_preview",    "TEXT DEFAULT NULL"),
        ("applications", "last_email_staging_id", "INTEGER DEFAULT NULL"),
        ("applications", "archived_round",        "TEXT DEFAULT NULL"),
        ("applications", "archived_at",           "TIMESTAMP DEFAULT NULL"),
        ("email_move_history", "move_source",     "TEXT DEFAULT 'manual'"),
    ]
    for table, col, definition in _migrations:
        try:
            with _conn() as db:
                db.execute(f"ALTER TABLE {table} ADD COLUMN {col} {definition}")
        except Exception:
            pass  # column already exists

    # Auto-backup once per calendar day (skip if DB doesn't exist yet)
    if _DB_PATH.exists():
        today = date.today().isoformat()
        try:
            with _conn() as _db:
                _row = _db.execute("SELECT value FROM settings WHERE key='last_backup_date'").fetchone()
                _last = _row["value"] if _row else ""
            if _last != today:
                backup_db()
                with _conn() as _db:
                    _db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('last_backup_date',?)", (today,))
        except Exception:
            pass  # never block startup over a backup failure

        # Auto-dismiss jobs older than 30 days, once per calendar day
        try:
            with _conn() as _db:
                _row = _db.execute("SELECT value FROM settings WHERE key='last_purge_old_date'").fetchone()
                _last = _row["value"] if _row else ""
            if _last != today:
                purge_old_jobs(max_age_days=30)
                with _conn() as _db:
                    _db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('last_purge_old_date',?)", (today,))
        except Exception:
            pass  # never block startup over a purge failure

        # Backfill has_easy_apply from applications table (safe to run repeatedly)
        try:
            backfilled = backfill_easy_apply_from_applications()
            if backfilled:
                log.info("Backfilled has_easy_apply=1 for %d previously-applied jobs", backfilled)
        except Exception:
            pass  # never block startup over a backfill failure

        # Purge dismissed jobs once per calendar day (clean up the DB)
        try:
            with _conn() as _db:
                _row = _db.execute("SELECT value FROM settings WHERE key='last_purge_dismissed_date'").fetchone()
                _last = _row["value"] if _row else ""
            if _last != today:
                purged = purge_dismissed_jobs()
                if purged:
                    log.info("Purged %d dismissed jobs from DB", purged)
                with _conn() as _db:
                    _db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('last_purge_dismissed_date',?)", (today,))
        except Exception:
            pass  # never block startup over a purge failure


# ── Settings ──────────────────────────────────────────────────────────────────

def get_setting(key: str, default: str = "") -> str:
    with _conn() as db:
        row = db.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with _conn() as db:
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (key, value))


# ── Excel import ──────────────────────────────────────────────────────────────

# Column name aliases from the existing tracker spreadsheet
_IMPORT_ALIASES: dict[str, list[str]] = {
    "company":          ["company", "employer"],
    "role":             ["role", "job title", "title", "position"],
    "location":         ["location", "city"],
    "date_applied":     ["date applied", "date", "application date"],
    "status":           ["status", "application status"],
    "match_pct":        ["match score", "match %", "score"],
    "job_url":          ["job url", "url", "link"],
    "source":           ["source"],
    "interview_chance": ["interview chance", "chance"],
    "archive_path":     ["archive folder", "archive", "folder"],
    "language":         ["german required", "german level", "german"],
}


def _alias_val(row: dict, col_lower: dict[str, str], key: str):
    for alias in _IMPORT_ALIASES.get(key, [key]):
        col = col_lower.get(alias.lower())
        if col is not None:
            v = row.get(col)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            if hasattr(v, "strftime"):
                return v.strftime("%Y-%m-%d")
            s = str(v).strip()
            if s:
                return s
    return None


def _parse_match_pct(raw) -> "int | None":
    if raw is None:
        return None
    try:
        v = float(str(raw).replace("%", "").strip())
        return round(v * 100) if 0 < v <= 1.0 else round(v)
    except (ValueError, TypeError):
        return None


def import_from_excel(filepath: "str | Path") -> int:
    """
    Read Applied / Interviews / Rejected sheets (or a single sheet with a
    Status column) and INSERT rows into applications.  Returns count inserted.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    init_db()
    with pd.ExcelFile(filepath) as xl:
        target_sheets = [s for s in ["Applied", "Interviews", "Rejected", "Offer"]
                         if s in xl.sheet_names]

        if not target_sheets:
            # Single-sheet workbook — infer status from Status column
            df = xl.parse(xl.sheet_names[0])
            df = df.where(pd.notna(df), other=None)
            return _import_df(df, status_override=None)

        total = 0
        for sheet in target_sheets:
            df = xl.parse(sheet)
            df = df.where(pd.notna(df), other=None)
            total += _import_df(df, status_override=sheet)
    return total


def _import_df(df: pd.DataFrame, status_override: "str | None") -> int:
    df.columns = [str(c).strip() for c in df.columns]
    col_lower = {c.lower(): c for c in df.columns}
    inserted = 0
    with _conn() as db:
        for _, row in df.iterrows():
            r = row.to_dict()
            company = _alias_val(r, col_lower, "company")
            role    = _alias_val(r, col_lower, "role")
            if not company and not role:
                continue  # empty row
            status = status_override or _alias_val(r, col_lower, "status") or "Applied"
            try:
                db.execute("""
                    INSERT OR IGNORE INTO applications
                    (company, role, location, date_applied, status, match_pct,
                     job_url, source, interview_chance, archive_path, language)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    company,
                    role,
                    _alias_val(r, col_lower, "location"),
                    _alias_val(r, col_lower, "date_applied"),
                    status,
                    _parse_match_pct(_alias_val(r, col_lower, "match_pct")),
                    _alias_val(r, col_lower, "job_url"),
                    _alias_val(r, col_lower, "source"),
                    _alias_val(r, col_lower, "interview_chance"),
                    _alias_val(r, col_lower, "archive_path"),
                    _alias_val(r, col_lower, "language"),
                ))
                inserted += db.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass
    return inserted


# ── Dashboard queries ─────────────────────────────────────────────────────────

def get_overview_data() -> dict:
    """Return stats + upcoming scheduled events + priority pending tasks."""
    with _conn() as db:
        status_rows = db.execute(
            "SELECT status, COUNT(*) AS cnt FROM applications WHERE archived_at IS NULL GROUP BY status"
        ).fetchall()

        upcoming = db.execute("""
            SELECT id, company, role, date_applied, status, job_url
            FROM applications
            WHERE archived_at IS NULL
              AND status IN (
                'Call Scheduled ✓',
                'Technical Scheduled ✓',
                'Final Scheduled ✓'
            )
            ORDER BY date_applied DESC
            LIMIT 10
        """).fetchall()

        tasks = db.execute("""
            SELECT id, company, role, date_applied, status, job_url
            FROM applications
            WHERE archived_at IS NULL
              AND status IN (
                '⏸️ Technical — awaiting your confirmation',
                '⏸️ Final — awaiting your confirmation',
                'Applied — unconfirmed'
            )
            ORDER BY date_applied ASC
        """).fetchall()

    buckets = {"Applied": 0, "Interviews": 0, "Rejected": 0, "Offer": 0}
    total = 0
    for r in status_rows:
        total += r["cnt"]
        sl = (r["status"] or "").lower()
        if "rejected" in sl:
            buckets["Rejected"] += r["cnt"]
        elif "offer" in sl:
            buckets["Offer"] += r["cnt"]
        elif any(k in sl for k in ("scheduled", "interview", "technical", "final", "call")):
            buckets["Interviews"] += r["cnt"]
        else:
            buckets["Applied"] += r["cnt"]

    auto_events = get_upcoming_events()
    return {
        "stats":           {"total": total, **buckets},
        "upcoming_events": [dict(r) for r in upcoming],
        "priority_tasks":  [dict(r) for r in tasks],
        "auto_events":     auto_events,
    }


def _status_bucket(status: str) -> str:
    """Map any status string to one of the four canonical display buckets."""
    sl = (status or "").lower()
    if "rejected" in sl:
        return "Rejected"
    if "offer" in sl:
        return "Offer"
    if any(k in sl for k in ("scheduled", "interview", "technical", "final", "call")):
        return "Interviews"
    return "Applied"


def get_dashboard_data() -> dict:
    with _conn() as db:
        rows = db.execute("""
            SELECT id, company, role, location, date_applied, status,
                   verdict, match_pct, key_gap, strengths, company_size,
                   language, job_url, source, interview_chance, archive_path,
                   COALESCE(applied_by, '')  AS applied_by,
                   COALESCE(apply_type, '')  AS apply_type,
                   last_email_date,
                   created_at
            FROM applications
            WHERE archived_at IS NULL
            ORDER BY date_applied DESC, created_at DESC
        """).fetchall()
        import_done = db.execute(
            "SELECT value FROM settings WHERE key='import_done'"
        ).fetchone()

    all_rows = [dict(r) for r in rows]
    statuses = ["Applied", "Interviews", "Rejected", "Offer"]
    grouped: dict[str, list] = {s: [] for s in statuses}
    for r in all_rows:
        bucket = _status_bucket(r.get("status", ""))
        grouped[bucket].append(r)

    overview = {"total": len(all_rows)}
    for s in statuses:
        overview[s] = len(grouped[s])

    return {
        "overview":    overview,
        "Applied":     grouped["Applied"],
        "Interviews":  grouped["Interviews"],
        "Rejected":    grouped["Rejected"],
        "Offer":       grouped["Offer"],
        "import_done": import_done is not None,
    }


def archive_all_applications(round_label: str = "") -> dict:
    """Archive every currently-active application under a round label.

    Archived rows are excluded from get_dashboard_data()/get_overview_data()
    but remain in the applications table (nothing is deleted), and stay
    visible via get_archived_data().
    """
    with _conn() as db:
        if not round_label:
            n = db.execute(
                "SELECT COUNT(DISTINCT archived_round) AS n FROM applications "
                "WHERE archived_round IS NOT NULL"
            ).fetchone()["n"]
            round_label = f"Round {n + 1} - {date.today().isoformat()}"
        db.execute(
            "UPDATE applications SET archived_round=?, archived_at=CURRENT_TIMESTAMP "
            "WHERE archived_at IS NULL",
            (round_label,),
        )
        archived = db.execute("SELECT changes()").fetchone()[0]
    return {"archived": archived, "round": round_label}


def get_archived_data() -> dict:
    """Return archived applications grouped by round, most recently archived first."""
    with _conn() as db:
        rows = db.execute("""
            SELECT id, company, role, location, date_applied, status,
                   verdict, match_pct, key_gap, strengths, company_size,
                   language, job_url, source, interview_chance, archive_path,
                   COALESCE(applied_by, '')  AS applied_by,
                   COALESCE(apply_type, '')  AS apply_type,
                   last_email_date, created_at,
                   archived_round, archived_at
            FROM applications
            WHERE archived_at IS NOT NULL
            ORDER BY archived_at DESC, date_applied DESC, created_at DESC
        """).fetchall()

    all_rows = [dict(r) for r in rows]
    rounds: dict[str, list] = {}
    order: list[str] = []
    for r in all_rows:
        label = r["archived_round"] or "Archived"
        if label not in rounds:
            rounds[label] = []
            order.append(label)
        rounds[label].append(r)

    return {
        "rounds": [{"label": label, "applications": rounds[label]} for label in order],
        "total":  len(all_rows),
    }


# ── Applications ──────────────────────────────────────────────────────────────

def get_applied_urls() -> set[str]:
    with _conn() as db:
        rows = db.execute(
            "SELECT job_url FROM applications WHERE job_url IS NOT NULL"
        ).fetchall()
    return {r["job_url"] for r in rows}


def is_already_applied(url: str) -> bool:
    if not url:
        return False
    with _conn() as db:
        row = db.execute(
            "SELECT id FROM applications WHERE job_url=?", (url,)
        ).fetchone()
    return row is not None


def exclude_job(url: str, company: str = "", title: str = "", reason: str = "") -> None:
    """Permanently exclude a job URL from future scraping/matching runs."""
    if not url:
        return
    with _conn() as db:
        db.execute("""
            INSERT INTO excluded_jobs (url, company, title, reason)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(url) DO UPDATE SET
                reason      = excluded.reason,
                excluded_at = CURRENT_TIMESTAMP
        """, (url.strip(), company or "", title or "", reason or ""))
        # Also mark dismissed in seen_jobs so the cache filter picks it up immediately
        db.execute("UPDATE seen_jobs SET dismissed=1 WHERE url=?", (url.strip(),))


def log_application(job: dict, archive_path: str = "", status: str = "Applied",
                    applied_by: str = "", apply_type: str = "") -> "int | None":
    url = (job.get("url") or "").strip() or None
    date_applied = (job.get("date_applied") or "").strip() or date.today().isoformat()
    with _conn() as db:
        if url:
            db.execute("""
                INSERT INTO applications
                (company, role, location, date_applied, status, match_pct,
                 job_url, source, interview_chance, archive_path, language,
                 applied_by, apply_type)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(job_url) DO UPDATE SET
                    status       = excluded.status,
                    archive_path = excluded.archive_path,
                    date_applied = excluded.date_applied,
                    applied_by   = excluded.applied_by,
                    apply_type   = excluded.apply_type
            """, (
                job.get("company", ""),
                job.get("title",   ""),
                job.get("location", ""),
                date_applied,
                status,
                job.get("match_score"),
                url,
                job.get("source", ""),
                job.get("interview_chance", ""),
                str(archive_path) if archive_path else "",
                job.get("german_level_required", ""),
                applied_by or "",
                apply_type or "",
            ))
            # lastrowid is unreliable on the upsert's UPDATE path — job_url is
            # unique, so look the row up directly rather than guess.
            row = db.execute("SELECT id FROM applications WHERE job_url=?", (url,)).fetchone()
            return row["id"] if row else None
        else:
            cur = db.execute("""
                INSERT INTO applications
                (company, role, location, date_applied, status, match_pct,
                 source, interview_chance, archive_path, language,
                 applied_by, apply_type)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                job.get("company", ""),
                job.get("title",   ""),
                job.get("location", ""),
                date_applied,
                status,
                job.get("match_score"),
                job.get("source", ""),
                job.get("interview_chance", ""),
                str(archive_path) if archive_path else "",
                job.get("german_level_required", ""),
                applied_by or "",
                apply_type or "",
            ))
            return cur.lastrowid


# ── Scraped / matched jobs ────────────────────────────────────────────────────

def insert_scraped_jobs(jobs: list[dict]) -> int:
    inserted = 0
    with _conn() as db:
        for job in jobs:
            url = (job.get("url") or "").strip()
            if not url:
                continue
            try:
                db.execute("""
                    INSERT OR IGNORE INTO scraped_jobs
                    (title, company, location, url, description, source, posted_date)
                    VALUES (?,?,?,?,?,?,?)
                """, (
                    job.get("title", ""),
                    job.get("company", ""),
                    job.get("location", ""),
                    url,
                    job.get("description", ""),
                    job.get("source", "LinkedIn"),
                    job.get("posted_date", ""),
                ))
                inserted += db.execute("SELECT changes()").fetchone()[0]
            except Exception:
                pass
    return inserted


def add_scraped_job(job: dict) -> bool:
    """Add a single scraped job. Returns True if inserted (new), False if duplicate."""
    return insert_scraped_jobs([job]) > 0


# ── seen_jobs cache ───────────────────────────────────────────────────────────

def _cutoff_for_config(cfg) -> str:
    """Return the ISO date string for the start of the configured time window."""
    from datetime import timedelta
    limit = (cfg or {}).get("posted_limit", "24h")
    now   = datetime.utcnow()
    delta = {"1h": timedelta(hours=1), "24h": timedelta(hours=24),
             "week": timedelta(weeks=1), "month": timedelta(days=30)}.get(limit, timedelta(hours=24))
    return (now - delta).strftime("%Y-%m-%d")


def get_cached_job_urls() -> set[str]:
    """Return all non-dismissed URLs from the persistent seen_jobs cache."""
    with _conn() as db:
        rows = db.execute("SELECT url FROM seen_jobs WHERE dismissed=0").fetchall()
        return {r["url"] for r in rows if r["url"]}


def upsert_seen_jobs(jobs: list[dict]) -> None:
    """Insert new jobs into the cache; update last_scraped_at for jobs seen again."""
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    with _conn() as db:
        for j in jobs:
            url = (j.get("url") or "").strip()
            if not url:
                continue
            pd = (j.get("posted_date") or "").strip()
            db.execute("""
                INSERT INTO seen_jobs (url, title, company, location, description, source,
                                       posted_date, first_scraped_at, last_scraped_at, has_easy_apply)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(url) DO UPDATE SET
                    last_scraped_at = excluded.last_scraped_at,
                    posted_date = CASE
                        WHEN excluded.posted_date != '' THEN excluded.posted_date
                        ELSE seen_jobs.posted_date
                    END,
                    has_easy_apply = excluded.has_easy_apply
            """, (
                url,
                j.get("title", ""),
                j.get("company", ""),
                j.get("location", ""),
                j.get("description", ""),
                j.get("source", "LinkedIn"),
                pd, now, now,
                1 if j.get("has_easy_apply") else 0,
            ))


def get_relevant_cached_jobs(cfg) -> list[dict]:
    """Return all cached jobs within the configured time window (scored + unscored)."""
    cutoff = _cutoff_for_config(cfg)
    with _conn() as db:
        rows = db.execute("""
            SELECT url, title, company, location, description, source,
                   posted_date, has_easy_apply,
                   COALESCE(first_scraped_at, first_seen_at) AS scraped_at,
                   match_score, interview_chance, skip_reason, german_level_required
            FROM seen_jobs
            WHERE dismissed = 0
              AND applied = 0
              AND (
                  posted_date = ''
                  OR posted_date IS NULL
                  OR posted_date >= ?
              )
              AND url NOT IN (
                  SELECT job_url FROM applications
                  WHERE job_url IS NOT NULL AND job_url != ''
              )
              AND url NOT IN (
                  SELECT url FROM excluded_jobs
                  WHERE url IS NOT NULL AND url != ''
              )
            ORDER BY posted_date DESC, COALESCE(first_scraped_at, first_seen_at) DESC
        """, (cutoff,)).fetchall()
        return [dict(r) for r in rows]


def get_all_cached_jobs() -> list[dict]:
    """Return every active (non-dismissed, non-applied) cached job, ignoring the
    posted-date time window — used by 'show all saved jobs' to recover jobs that
    scrolled outside the window before ever being matched."""
    with _conn() as db:
        rows = db.execute("""
            SELECT url, title, company, location, description, source,
                   posted_date, has_easy_apply,
                   COALESCE(first_scraped_at, first_seen_at) AS scraped_at,
                   match_score, interview_chance, skip_reason, german_level_required
            FROM seen_jobs
            WHERE dismissed = 0
              AND applied = 0
              AND url NOT IN (
                  SELECT job_url FROM applications
                  WHERE job_url IS NOT NULL AND job_url != ''
              )
              AND url NOT IN (
                  SELECT url FROM excluded_jobs
                  WHERE url IS NOT NULL AND url != ''
              )
            ORDER BY posted_date DESC, COALESCE(first_scraped_at, first_seen_at) DESC
        """).fetchall()
        return [dict(r) for r in rows]


def get_all_scored_jobs() -> list[dict]:
    """Return every active job that already has a match_score from a previous
    matching run — used by 'show already matched' so re-running the matcher
    (which only scores unscored jobs) isn't the only way to see past results."""
    with _conn() as db:
        rows = db.execute("""
            SELECT url, title, company, location, description, source,
                   posted_date, has_easy_apply,
                   COALESCE(first_scraped_at, first_seen_at) AS scraped_at,
                   match_score, interview_chance, skip_reason, german_level_required
            FROM seen_jobs
            WHERE dismissed = 0
              AND applied = 0
              AND match_score IS NOT NULL
              AND url NOT IN (
                  SELECT job_url FROM applications
                  WHERE job_url IS NOT NULL AND job_url != ''
              )
              AND url NOT IN (
                  SELECT url FROM excluded_jobs
                  WHERE url IS NOT NULL AND url != ''
              )
            ORDER BY match_score DESC
        """).fetchall()
        return [dict(r) for r in rows]


def bulk_populate_matched_jobs(jobs: list[dict]) -> int:
    """After scraped_jobs has been (re)populated from cache, create matched_jobs
    rows for any of those jobs that already carry a match_score — keyed by URL
    since the cache has no live scraped_jobs.id of its own."""
    count = 0
    with _conn() as db:
        for j in jobs:
            if j.get("match_score") is None:
                continue
            url = (j.get("url") or "").strip()
            if not url:
                continue
            row = db.execute("SELECT id FROM scraped_jobs WHERE url=?", (url,)).fetchone()
            if not row:
                continue
            db.execute("""
                INSERT INTO matched_jobs
                (scraped_job_id, match_score, interview_chance, german_level, skip_reason, match_summary)
                VALUES (?,?,?,?,?,?)
            """, (
                row["id"],
                j.get("match_score", 0),
                j.get("interview_chance", "low"),
                j.get("german_level_required", "none"),
                j.get("skip_reason"),
                j.get("match_summary", ""),
            ))
            count += 1
    return count


def get_cache_window_stats(cfg) -> dict:
    """Return counts: jobs within the time window, outside it, and total active."""
    cutoff = _cutoff_for_config(cfg)
    with _conn() as db:
        within = db.execute("""
            SELECT COUNT(*) FROM seen_jobs
            WHERE dismissed = 0 AND applied = 0
              AND (
                  posted_date = ''
                  OR posted_date IS NULL
                  OR posted_date >= ?
              )
              AND url NOT IN (
                  SELECT job_url FROM applications WHERE job_url IS NOT NULL AND job_url != ''
              )
        """, (cutoff,)).fetchone()[0]

        total_active = db.execute("""
            SELECT COUNT(*) FROM seen_jobs
            WHERE dismissed = 0 AND applied = 0
              AND url NOT IN (
                  SELECT job_url FROM applications WHERE job_url IS NOT NULL AND job_url != ''
              )
        """).fetchone()[0]

    return {
        "within_window":  within,
        "outside_window": max(0, total_active - within),
        "total_active":   total_active,
    }


def update_seen_job_score(url: str, scores: dict, resume_hash: str = "") -> None:
    """Store Claude's scoring result back into the seen_jobs cache."""
    with _conn() as db:
        db.execute("""
            UPDATE seen_jobs
            SET match_score=?, interview_chance=?, skip_reason=?,
                german_level_required=?, resume_hash=?
            WHERE url=?
        """, (
            scores.get("match_score"),
            scores.get("interview_chance"),
            scores.get("skip_reason"),
            scores.get("german_level_required"),
            resume_hash or None,
            url,
        ))


def dismiss_job(url: str) -> None:
    """Mark a job as dismissed — never shown again."""
    with _conn() as db:
        db.execute("UPDATE seen_jobs SET dismissed=1 WHERE url=?", (url,))


def purge_job_by_url(url: str, reason: str = "dismissed") -> None:
    """Permanently remove a job across all tables so it never resurfaces.

    Actions (all in one transaction):
    - INSERT into excluded_jobs → blocks stage-3 cache restoration on future scrapes
    - seen_jobs.dismissed = 1  → blocks fresh-scrape dedup path
    - DELETE from matched_jobs → cleans current match results
    - DELETE from scraped_jobs → cleans current review list
    """
    url = (url or "").strip()
    if not url:
        return
    with _conn() as db:
        row = db.execute(
            "SELECT company, title FROM seen_jobs WHERE url=? LIMIT 1", (url,)
        ).fetchone()
        company = row["company"] if row else ""
        title   = row["title"]   if row else ""
        db.execute("""
            INSERT INTO excluded_jobs (url, company, title, reason)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(url) DO UPDATE SET
                reason      = excluded.reason,
                excluded_at = CURRENT_TIMESTAMP
        """, (url, company or "", title or "", reason))
        db.execute("UPDATE seen_jobs SET dismissed=1 WHERE url=?", (url,))
        db.execute("""
            DELETE FROM matched_jobs
            WHERE scraped_job_id IN (SELECT id FROM scraped_jobs WHERE url=?)
        """, (url,))
        db.execute("DELETE FROM scraped_jobs WHERE url=?", (url,))


def purge_low_score_jobs(min_score: int = 50) -> int:
    """Auto-dismiss all jobs with match_score < min_score.

    For each qualifying job:
    - Inserts into excluded_jobs (permanent scrape block, survives table clears)
    - Sets seen_jobs.dismissed = 1 (blocks fresh-scrape and cache-restore paths)
    - Deletes from matched_jobs and scraped_jobs (cleans the current run)

    Sweeps seen_jobs directly (not just the current run's matched_jobs/scraped_jobs
    join) — those two tables get wiped on every scrape, so a job scored low in an
    earlier run before this function ran (or before it existed) would otherwise
    leak permanently as an un-dismissed low score. seen_jobs.match_score is the
    durable source of truth; matched_jobs/scraped_jobs cleanup below is just for
    the current run's display.

    Returns the count of purged jobs.
    """
    with _conn() as db:
        rows = db.execute("""
            SELECT url, company, title, match_score
            FROM   seen_jobs
            WHERE  match_score IS NOT NULL AND match_score < ? AND dismissed = 0
        """, (min_score,)).fetchall()

        if not rows:
            return 0

        urls = []
        for r in rows:
            url = (r["url"] or "").strip()
            if not url:
                continue
            urls.append(url)
            db.execute("""
                INSERT INTO excluded_jobs (url, company, title, reason)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(url) DO UPDATE SET
                    reason      = excluded.reason,
                    excluded_at = CURRENT_TIMESTAMP
            """, (url, r["company"] or "", r["title"] or "",
                  f"auto-dismissed: low score ({r['match_score']}%)"))
            db.execute("UPDATE seen_jobs SET dismissed=1 WHERE url=?", (url,))

        if urls:
            ph = ",".join("?" * len(urls))
            db.execute(f"""
                DELETE FROM matched_jobs WHERE scraped_job_id IN (
                    SELECT id FROM scraped_jobs WHERE url IN ({ph})
                )
            """, urls)
            db.execute(f"DELETE FROM scraped_jobs WHERE url IN ({ph})", urls)

    return len(rows)


def purge_old_jobs(max_age_days: int = 30) -> int:
    """Auto-dismiss every active job older than max_age_days, scored or not.

    Age is COALESCE(posted_date, first_scraped_at, first_seen_at) — most
    LinkedIn jobs have posted_date, but some don't, so this falls back to when
    we first saw it rather than silently never expiring those.

    Same effect as purge_low_score_jobs: excluded_jobs (permanent scrape
    block) + seen_jobs.dismissed=1 + cleanup of the current run's
    matched_jobs/scraped_jobs rows.
    """
    from datetime import timedelta
    cutoff = (datetime.utcnow() - timedelta(days=max_age_days)).strftime("%Y-%m-%d")
    with _conn() as db:
        rows = db.execute("""
            SELECT url, company, title,
                   COALESCE(posted_date, first_scraped_at, first_seen_at) AS age_date
            FROM   seen_jobs
            WHERE  dismissed = 0
              AND  COALESCE(posted_date, first_scraped_at, first_seen_at) < ?
        """, (cutoff,)).fetchall()

        if not rows:
            return 0

        urls = []
        for r in rows:
            url = (r["url"] or "").strip()
            if not url:
                continue
            urls.append(url)
            db.execute("""
                INSERT INTO excluded_jobs (url, company, title, reason)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(url) DO UPDATE SET
                    reason      = excluded.reason,
                    excluded_at = CURRENT_TIMESTAMP
            """, (url, r["company"] or "", r["title"] or "",
                  f"auto-dismissed: older than {max_age_days} days ({r['age_date']})"))
            db.execute("UPDATE seen_jobs SET dismissed=1 WHERE url=?", (url,))

        if urls:
            ph = ",".join("?" * len(urls))
            db.execute(f"""
                DELETE FROM matched_jobs WHERE scraped_job_id IN (
                    SELECT id FROM scraped_jobs WHERE url IN ({ph})
                )
            """, urls)
            db.execute(f"DELETE FROM scraped_jobs WHERE url IN ({ph})", urls)

    return len(rows)


def dismiss_job_by_id(job_id: int) -> None:
    with _conn() as db:
        db.execute("UPDATE seen_jobs SET dismissed=1 WHERE id=?", (job_id,))


def undismiss_job_by_id(job_id: int) -> None:
    with _conn() as db:
        db.execute("UPDATE seen_jobs SET dismissed=0 WHERE id=?", (job_id,))


def purge_dismissed_jobs() -> int:
    """Delete all dismissed jobs entirely from seen_jobs (they're blocked from scrapes
    via excluded_jobs anyway, so keeping them around just clutters the DB)."""
    with _conn() as db:
        count = db.execute("DELETE FROM seen_jobs WHERE dismissed=1").rowcount
    return count


def backfill_easy_apply_from_applications() -> int:
    """For jobs that were successfully applied to via LinkedIn Easy Apply,
    retroactively tag them as has_easy_apply=1 in seen_jobs. This recovers
    the tag for cached jobs that lost it during schema updates."""
    count = 0
    with _conn() as db:
        rows = db.execute("""
            SELECT DISTINCT job_url FROM applications
            WHERE job_url IS NOT NULL
              AND job_url != ''
              AND apply_type LIKE '%easy%'
        """).fetchall()

        for r in rows:
            url = (r["job_url"] or "").strip()
            if not url:
                continue
            result = db.execute(
                "UPDATE seen_jobs SET has_easy_apply=1 WHERE url=? AND has_easy_apply=0",
                (url,)
            )
            if result.rowcount > 0:
                count += 1

    return count


def get_cache_stats() -> dict:
    """Return aggregate counts for the seen_jobs table."""
    with _conn() as db:
        total     = db.execute("SELECT COUNT(*) FROM seen_jobs").fetchone()[0]
        unseen    = db.execute(
            "SELECT COUNT(*) FROM seen_jobs WHERE dismissed=0 AND match_score IS NULL"
        ).fetchone()[0]
        applied   = db.execute("SELECT COUNT(*) FROM seen_jobs WHERE applied=1").fetchone()[0]
        dismissed = db.execute("SELECT COUNT(*) FROM seen_jobs WHERE dismissed=1").fetchone()[0]
    return {"total_cached": total, "unseen": unseen, "applied": applied, "dismissed": dismissed}


# ── Scoring cache with resume hash ────────────────────────────────────────────

def get_resume_hash(cfg: dict) -> str:
    """Return the MD5 hex digest of the resume PDF, or '' if not found."""
    import hashlib
    resume_path = cfg["paths"]["resume_en"]
    p = Path(resume_path)
    if not p.exists():
        return ""
    h = hashlib.md5()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def get_scored_urls_for_hash(current_hash: str) -> set:
    """Return URLs already scored with the given resume hash."""
    if not current_hash:
        return set()
    with _conn() as db:
        rows = db.execute(
            "SELECT url FROM seen_jobs WHERE match_score IS NOT NULL AND resume_hash = ?",
            (current_hash,),
        ).fetchall()
        return {r["url"] for r in rows if r["url"]}


def has_scores_with_different_hash(current_hash: str, cfg=None) -> bool:
    """Return True if any seen_job within the time window has a stale or missing resume hash."""
    if not current_hash:
        return False
    cutoff = _cutoff_for_config(cfg)
    with _conn() as db:
        row = db.execute(
            "SELECT COUNT(*) FROM seen_jobs "
            "WHERE match_score IS NOT NULL "
            "  AND (resume_hash IS NULL OR resume_hash != ?) "
            "  AND applied = 0 AND dismissed = 0 "
            "  AND (posted_date = '' OR posted_date IS NULL OR posted_date >= ?)",
            (current_hash, cutoff),
        ).fetchone()
        return row[0] > 0


def invalidate_scores_for_changed_resume(cfg=None) -> int:
    """Clear cached scores for jobs within the configured time window; return row count."""
    cutoff = _cutoff_for_config(cfg)
    with _conn() as db:
        count = db.execute(
            "SELECT COUNT(*) FROM seen_jobs "
            "WHERE match_score IS NOT NULL AND applied=0 AND dismissed=0 "
            "  AND (posted_date = '' OR posted_date IS NULL OR posted_date >= ?)",
            (cutoff,),
        ).fetchone()[0]
        db.execute(
            "UPDATE seen_jobs "
            "SET match_score=NULL, interview_chance=NULL, skip_reason=NULL, "
            "    german_level_required=NULL, resume_hash=NULL "
            "WHERE match_score IS NOT NULL AND applied=0 AND dismissed=0 "
            "  AND (posted_date = '' OR posted_date IS NULL OR posted_date >= ?)",
            (cutoff,),
        )
        return count


def get_cached_scores_by_url(urls) -> dict:
    """Return a {url: scores_dict} map for the given URL set."""
    if not urls:
        return {}
    placeholders = ",".join("?" * len(urls))
    with _conn() as db:
        rows = db.execute(
            f"SELECT url, match_score, interview_chance, skip_reason, german_level_required "
            f"FROM seen_jobs WHERE url IN ({placeholders})",
            list(urls),
        ).fetchall()
        return {r["url"]: dict(r) for r in rows}


def build_score_cache(resume_hash: str) -> dict:
    """Build a comprehensive score cache keyed by URL and (company, title) tuple.

    Returns a dict where both url strings and (company_lower, title_lower) tuples
    map to a scores dict — enabling fast O(1) cache lookup without per-job DB queries.
    """
    if not resume_hash:
        return {}
    with _conn() as db:
        rows = db.execute("""
            SELECT url, company, title, match_score, interview_chance,
                   skip_reason, german_level_required, match_summary
            FROM seen_jobs
            WHERE resume_hash = ?
              AND match_score IS NOT NULL
              AND dismissed = 0
        """, (resume_hash,)).fetchall()
    cache: dict = {}
    for row in rows:
        d = dict(row)
        if d.get("url"):
            cache[d["url"]] = d
        co = (d.get("company") or "").lower().strip()
        ti = (d.get("title") or "").lower().strip()
        if co and ti:
            cache[(co, ti)] = d
    return cache


def get_unmatched_scraped_jobs() -> list[dict]:
    with _conn() as db:
        rows = db.execute("""
            SELECT s.id, s.title, s.company, s.location, s.url, s.description, s.source
            FROM scraped_jobs s
            LEFT JOIN matched_jobs m ON m.scraped_job_id = s.id
            WHERE m.id IS NULL
        """).fetchall()
    return [dict(r) for r in rows]


def insert_matched_job(scraped_job_id: int, scores: dict) -> None:
    with _conn() as db:
        db.execute("""
            INSERT INTO matched_jobs
            (scraped_job_id, match_score, interview_chance,
             german_level, skip_reason, match_summary)
            VALUES (?,?,?,?,?,?)
        """, (
            scraped_job_id,
            scores.get("match_score", 0),
            scores.get("interview_chance", "low"),
            scores.get("german_level_required", "none"),
            scores.get("skip_reason"),
            scores.get("match_summary", ""),
        ))


def get_matched_jobs_for_apply(cfg: dict) -> list[dict]:
    min_score   = cfg.get("min_match_score", 70)
    skip_levels = {lvl.lower() for lvl in cfg.get("skip_german_levels", [])}
    with _conn() as db:
        rows = db.execute("""
            SELECT s.id  AS scraped_id,
                   s.title, s.company, s.location, s.url, s.description, s.source,
                   m.match_score, m.interview_chance,
                   m.german_level, m.skip_reason, m.match_summary
            FROM matched_jobs m
            JOIN scraped_jobs s ON s.id = m.scraped_job_id
            WHERE m.match_score >= ?
              AND m.skip_reason IS NULL
              AND (s.url IS NULL OR s.url NOT IN
                   (SELECT job_url FROM applications WHERE job_url IS NOT NULL))
            ORDER BY m.match_score DESC
        """, (min_score,)).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        if (d.get("german_level") or "none").lower() in skip_levels:
            continue
        result.append({
            "id":                    d["scraped_id"],
            "scraped_id":             d["scraped_id"],
            "title":                 d["title"],
            "company":               d["company"],
            "location":              d["location"],
            "url":                   d["url"],
            "description":           d["description"],
            "source":                d["source"],
            "match_score":           d["match_score"],
            "interview_chance":      d["interview_chance"],
            "german_level_required": d["german_level"],
            "match_summary":         d["match_summary"],
        })
    return result


# ── Excel export ──────────────────────────────────────────────────────────────

def generate_excel_report() -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    with _conn() as db:
        rows = db.execute("""
            SELECT company, role, location, date_applied, status, match_pct,
                   interview_chance, language, job_url, archive_path, source
            FROM applications ORDER BY date_applied DESC, created_at DESC
        """).fetchall()

    all_rows = [dict(r) for r in rows]
    HEADERS = ["Company", "Role", "Location", "Date Applied", "Status",
               "Match %", "Interview Chance", "German", "Job URL", "Archive", "Source"]
    COL_WIDTHS = [20, 28, 16, 13, 14, 9, 16, 10, 45, 45, 12]

    def to_row(r):
        pct = r["match_pct"]
        return [
            r["company"], r["role"], r["location"], r["date_applied"], r["status"],
            f"{pct}%" if pct is not None else "",
            r["interview_chance"], r["language"], r["job_url"], r["archive_path"], r["source"],
        ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    grouped = {s: [r for r in all_rows if r["status"] == s]
               for s in ["Applied", "Interviews", "Rejected", "Offer"]}

    for sheet_name, sheet_rows in [
        ("Overview",   all_rows),
        ("Applied",    grouped["Applied"]),
        ("Interviews", grouped["Interviews"]),
        ("Rejected",   grouped["Rejected"]),
        ("Offer",      grouped["Offer"]),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        for i, cell in enumerate(ws[1], 0):
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(i + 1)].width = COL_WIDTHS[i]
        for r in sheet_rows:
            ws.append(to_row(r))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Upcoming interviews (manual) ──────────────────────────────────────────────

def get_upcoming_interviews() -> list[dict]:
    with _conn() as db:
        rows = db.execute(
            "SELECT * FROM upcoming_interviews ORDER BY date ASC, created_at ASC"
        ).fetchall()
    return [dict(r) for r in rows]


def add_interview(data: dict) -> int:
    with _conn() as db:
        cur = db.execute("""
            INSERT INTO upcoming_interviews
            (date, company, role, interview_type, time_berlin, format, job_url, notes)
            VALUES (?,?,?,?,?,?,?,?)
        """, (
            data.get("date", ""),
            data.get("company", ""),
            data.get("role", ""),
            data.get("interview_type", ""),
            data.get("time_berlin", ""),
            data.get("format", ""),
            data.get("job_url", ""),
            data.get("notes", ""),
        ))
        return cur.lastrowid


def update_interview(interview_id: int, data: dict) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE upcoming_interviews
            SET date=?, company=?, role=?, interview_type=?,
                time_berlin=?, format=?, job_url=?, notes=?
            WHERE id=?
        """, (
            data.get("date", ""),
            data.get("company", ""),
            data.get("role", ""),
            data.get("interview_type", ""),
            data.get("time_berlin", ""),
            data.get("format", ""),
            data.get("job_url", ""),
            data.get("notes", ""),
            interview_id,
        ))


def delete_interview(interview_id: int) -> None:
    with _conn() as db:
        db.execute("DELETE FROM upcoming_interviews WHERE id=?", (interview_id,))


# ── Priority tasks (manual) ────────────────────────────────────────────────────

def get_priority_tasks_manual() -> list[dict]:
    with _conn() as db:
        rows = db.execute(
            "SELECT * FROM priority_tasks ORDER BY created_at ASC"
        ).fetchall()
    return [dict(r) for r in rows]


def add_task_manual(data: dict) -> int:
    with _conn() as db:
        cur = db.execute("""
            INSERT INTO priority_tasks
            (priority, company, action, deadline, status, notes)
            VALUES (?,?,?,?,?,?)
        """, (
            data.get("priority", "NORMAL"),
            data.get("company", ""),
            data.get("action", ""),
            data.get("deadline", ""),
            data.get("status", ""),
            data.get("notes", ""),
        ))
        return cur.lastrowid


def update_task_manual(task_id: int, data: dict) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE priority_tasks
            SET priority=?, company=?, action=?, deadline=?, status=?, notes=?
            WHERE id=?
        """, (
            data.get("priority", "NORMAL"),
            data.get("company", ""),
            data.get("action", ""),
            data.get("deadline", ""),
            data.get("status", ""),
            data.get("notes", ""),
            task_id,
        ))


def delete_task_manual(task_id: int) -> None:
    with _conn() as db:
        db.execute("DELETE FROM priority_tasks WHERE id=?", (task_id,))


# ── Manual apply queue ────────────────────────────────────────────────────────

def log_manual_apply(job_url: str, title: str, company: str, platform: str,
                     note: str, screenshot_path: str = "", session_id: str = "") -> int:
    with _conn() as db:
        cur = db.execute("""
            INSERT INTO manual_apply_queue
            (job_url, title, company, platform, note, screenshot_path, session_id)
            VALUES (?,?,?,?,?,?,?)
        """, (job_url or "", title or "", company or "", platform or "",
              note or "", screenshot_path or "", session_id or ""))
        return cur.lastrowid


def get_latest_manual_session_id() -> str | None:
    with _conn() as db:
        row = db.execute(
            "SELECT session_id FROM manual_apply_queue "
            "WHERE session_id IS NOT NULL AND session_id != '' "
            "ORDER BY id DESC LIMIT 1"
        ).fetchone()
    return row[0] if row else None


def get_manual_queue(status: str = "pending", session_id: str | None = None) -> list[dict]:
    applied_filter = (
        " AND (job_url IS NULL OR job_url NOT IN "
        "(SELECT job_url FROM applications WHERE job_url IS NOT NULL))"
    )
    with _conn() as db:
        if session_id:
            rows = db.execute(
                "SELECT * FROM manual_apply_queue WHERE status=? AND session_id=?"
                + applied_filter + " ORDER BY created_at DESC",
                (status, session_id),
            ).fetchall()
        else:
            rows = db.execute(
                "SELECT * FROM manual_apply_queue WHERE status=?"
                + applied_filter + " ORDER BY created_at DESC",
                (status,),
            ).fetchall()
    return [dict(r) for r in rows]


def update_manual_queue_status(entry_id: int, status: str) -> None:
    with _conn() as db:
        db.execute("UPDATE manual_apply_queue SET status=? WHERE id=?", (status, entry_id))


# ── Apply sessions ────────────────────────────────────────────────────────────

def create_apply_session(total_jobs: int) -> int:
    with _conn() as db:
        cur = db.execute(
            "INSERT INTO apply_sessions (total_jobs) VALUES (?)", (total_jobs,)
        )
        return cur.lastrowid


def update_apply_session(session_id: int, finished_at: str,
                         success_count: int, manual_count: int,
                         failed_count: int) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE apply_sessions
            SET finished_at=?, success_count=?, manual_count=?, failed_count=?
            WHERE id=?
        """, (finished_at, success_count, manual_count, failed_count, session_id))


# ── Helper functions ──────────────────────────────────────────────────────────

def count_applications() -> int:
    """Return total count of applications in the database."""
    with _conn() as db:
        return db.execute("SELECT COUNT(*) FROM applications").fetchone()[0]


def get_all_jobs() -> list[dict]:
    """Return all applications from the database."""
    with _conn() as db:
        rows = db.execute("""
            SELECT id, company, role, location, date_applied, status,
                   verdict, match_pct, key_gap, strengths, company_size,
                   language, job_url, source, interview_chance, archive_path
            FROM applications
            ORDER BY date_applied DESC, created_at DESC
        """).fetchall()
    return [dict(r) for r in rows]


# ── Email staging ──────────────────────────────────────────────────────────────

def stage_email(record: dict) -> int:
    """Insert a new email staging record. Returns new id, or 0 on duplicate."""
    with _conn() as db:
        cur = db.execute("""
            INSERT OR IGNORE INTO email_staging
            (email_uid, email_message_id, sender, subject, body_preview,
             received_date, source_folder, matched_app_id, match_confidence,
             match_type, predicted_folder, confidence_score,
             classification_reason, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            record.get("email_uid", ""),
            record.get("email_message_id", ""),
            record.get("sender", ""),
            record.get("subject", ""),
            record.get("body_preview", ""),
            record.get("received_date", ""),
            record.get("source_folder", ""),
            record.get("matched_app_id"),
            record.get("match_confidence", 0),
            record.get("match_type", "unmatched"),
            record.get("predicted_folder", "Uncertain"),
            record.get("confidence_score", 0),
            record.get("classification_reason", ""),
            "pending",
        ))
        return cur.lastrowid


def get_staged_email(email_id: int) -> "dict | None":
    with _conn() as db:
        row = db.execute(
            "SELECT * FROM email_staging WHERE id=?", (email_id,)
        ).fetchone()
    return dict(row) if row else None


def get_pending_emails() -> list[dict]:
    with _conn() as db:
        rows = db.execute("""
            SELECT es.*, a.company, a.role
            FROM email_staging es
            LEFT JOIN applications a ON a.id = es.matched_app_id
            WHERE es.status = 'pending'
            ORDER BY es.created_at DESC
        """).fetchall()
    return [dict(r) for r in rows]


def get_skipped_emails() -> list[dict]:
    with _conn() as db:
        rows = db.execute("""
            SELECT es.*, a.company, a.role
            FROM email_staging es
            LEFT JOIN applications a ON a.id = es.matched_app_id
            WHERE es.status = 'skipped'
            ORDER BY es.created_at DESC
        """).fetchall()
    return [dict(r) for r in rows]


def requeue_email(email_id: int) -> None:
    """Reset a skipped email back to pending so it re-appears for approval."""
    with _conn() as db:
        db.execute(
            "UPDATE email_staging SET status='pending', reviewed_at=NULL WHERE id=?",
            (email_id,),
        )


def get_failed_email_count() -> int:
    """Return how many email_staging rows are in 'failed' status."""
    with _conn() as db:
        row = db.execute(
            "SELECT COUNT(*) AS n FROM email_staging WHERE status = 'failed'"
        ).fetchone()
    return row["n"] if row else 0


def restage_falsely_executed(live_message_ids: set) -> int:
    """Re-stage executed emails that are still visible in monitored Outlook folders.

    If an email is marked 'executed' but its internetMessageId still appears in a
    monitored folder (Inbox/Focus/Other/Junk), the Graph move never actually
    happened — reset it to 'pending' so it shows up in the approval queue again.
    Returns the count of rows updated.
    """
    if not live_message_ids:
        return 0
    with _conn() as db:
        ph = ",".join("?" * len(live_message_ids))
        rows = db.execute(
            f"SELECT id FROM email_staging "
            f"WHERE status = 'executed' AND email_message_id IN ({ph})",
            list(live_message_ids),
        ).fetchall()
        if not rows:
            return 0
        ids = [r["id"] for r in rows]
        id_ph = ",".join("?" * len(ids))
        db.execute(
            f"UPDATE email_staging SET status='pending', executed_at=NULL "
            f"WHERE id IN ({id_ph})",
            ids,
        )
    return len(ids)


def get_staged_message_ids() -> set[str]:
    """Return Message-IDs that should be skipped on the next processor scan.

    Failed emails are excluded so they can be re-staged and retried.
    Pending/executed/skipped emails are excluded to prevent duplicates.
    """
    with _conn() as db:
        rows = db.execute(
            "SELECT email_message_id FROM email_staging "
            "WHERE email_message_id IS NOT NULL AND status != 'failed'"
        ).fetchall()
    return {r["email_message_id"] for r in rows}


def update_email_uid(email_id: int, new_graph_id: str) -> None:
    """Update the stored Graph message ID (email_uid) after it changes due to a folder move."""
    with _conn() as db:
        db.execute("UPDATE email_staging SET email_uid=? WHERE id=?", (new_graph_id, email_id))


def update_email_staging_status(email_id: int, status: str,
                                  executed_at: "str | None" = None,
                                  reviewed_at: "str | None" = None) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE email_staging
            SET status=?,
                executed_at=COALESCE(?, executed_at),
                reviewed_at=COALESCE(?, reviewed_at)
            WHERE id=?
        """, (status, executed_at, reviewed_at, email_id))


def set_email_override_folder(email_id: int, folder: str) -> None:
    with _conn() as db:
        db.execute(
            "UPDATE email_staging SET user_override_folder=? WHERE id=?",
            (folder, email_id),
        )


def get_email_logs(limit: int = 100, offset: int = 0) -> list[dict]:
    with _conn() as db:
        rows = db.execute("""
            SELECT h.*, es.sender, es.subject, es.predicted_folder
            FROM email_move_history h
            LEFT JOIN email_staging es ON es.id = h.email_staging_id
            ORDER BY h.moved_at DESC
            LIMIT ? OFFSET ?
        """, (limit, offset)).fetchall()
    return [dict(r) for r in rows]


def delete_email_move_log(staging_id: int) -> None:
    """Remove all log entries for a staging record (used when a move is undone via Restore)."""
    with _conn() as db:
        db.execute("DELETE FROM email_move_history WHERE email_staging_id=?", (staging_id,))


def log_email_move(staging_id: int, from_folder: str, to_folder: str,
                   success: bool, error: "str | None" = None,
                   move_source: str = "manual") -> None:
    with _conn() as db:
        db.execute("""
            INSERT INTO email_move_history
            (email_staging_id, from_folder, to_folder, success, error_message, move_source)
            VALUES (?,?,?,?,?,?)
        """, (staging_id, from_folder, to_folder, int(success), error or "", move_source))


def update_application_from_email(app_id: int, status: str,
                                   email_date: str, preview: str,
                                   staging_id: int) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE applications
            SET status=?, last_email_date=?, last_email_preview=?,
                last_email_staging_id=?
            WHERE id=?
        """, (status, email_date, preview[:200], staging_id, app_id))


def get_application_companies() -> list[dict]:
    """Return id + company for all applications, used for email-to-job matching."""
    with _conn() as db:
        rows = db.execute(
            "SELECT id, company FROM applications "
            "WHERE company IS NOT NULL AND company != '' "
            "ORDER BY id DESC"
        ).fetchall()
    return [dict(r) for r in rows]


def get_application_by_id(app_id: int) -> "dict | None":
    """Return the full applications row for app_id, or None."""
    with _conn() as db:
        row = db.execute("SELECT * FROM applications WHERE id=?", (app_id,)).fetchone()
    return dict(row) if row else None


def get_event_by_source_email(staging_id: int) -> "dict | None":
    """Return the upcoming_event row that was created from a specific staging email."""
    with _conn() as db:
        row = db.execute(
            "SELECT * FROM upcoming_events WHERE source_email_id=? ORDER BY id DESC LIMIT 1",
            (staging_id,),
        ).fetchone()
    return dict(row) if row else None


def upsert_event_from_executor(staging_id: int, app_id: int, data: dict,
                               received_date: str = "") -> None:
    """
    Create or update the upcoming_event for this application.

    Lookup is by app_id (not source_email_id) so that all emails about the same
    application converge on a single event row.

    Date priority: the incoming email's received_date is compared to the
    received_date of whichever email last updated the event.
    - Newer email → overwrite all provided fields (reschedule notices win).
    - Older email → skip to avoid clobbering more-current data.
    - If no date info available → always update (safe fallback).
    """
    with _conn() as conn:
        # Find the most recent scheduled event for this application
        row = conn.execute("""
            SELECT ue.*, COALESCE(es.received_date, '') AS src_date
            FROM upcoming_events ue
            LEFT JOIN email_staging es ON es.id = ue.source_email_id
            WHERE ue.app_id = ? AND ue.status = 'scheduled'
            ORDER BY ue.id DESC LIMIT 1
        """, (app_id,)).fetchone()

        if row:
            existing_src_date = row["src_date"] or ""
            # Allow update if incoming is newer OR we can't compare dates
            should_update = (not existing_src_date or not received_date
                             or received_date >= existing_src_date)
            if should_update:
                conn.execute("""
                    UPDATE upcoming_events SET
                        source_email_id = ?,
                        event_type  = COALESCE(NULLIF(?, ''), event_type),
                        title       = COALESCE(NULLIF(?, ''), title),
                        description = COALESCE(NULLIF(?, ''), description),
                        event_date  = COALESCE(NULLIF(?, ''), event_date),
                        event_time  = COALESCE(NULLIF(?, ''), event_time),
                        timezone    = COALESCE(NULLIF(?, ''), timezone),
                        priority    = ?
                    WHERE id = ?
                """, (
                    staging_id,
                    data.get("event_type") or None,
                    data.get("title") or None,
                    data.get("description") or None,
                    data.get("event_date") or None,
                    data.get("event_time") or None,
                    data.get("timezone") or None,
                    data.get("priority", "high"),
                    row["id"],
                ))
        else:
            conn.execute("""
                INSERT INTO upcoming_events
                    (app_id, event_type, title, description, event_date,
                     event_time, timezone, priority, source_email_id, status)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                app_id,
                data.get("event_type", "interview"),
                data.get("title", ""),
                data.get("description", ""),
                data.get("event_date"),
                data.get("event_time"),
                data.get("timezone"),
                data.get("priority", "high"),
                staging_id,
                "scheduled",
            ))


def insert_upcoming_event(record: dict) -> int:
    with _conn() as db:
        cur = db.execute("""
            INSERT INTO upcoming_events
            (app_id, event_type, title, description, event_date, event_time,
             timezone, priority, source_email_id, status)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            record.get("app_id"),
            record.get("event_type", "interview"),
            record.get("title", ""),
            record.get("description", ""),
            record.get("event_date"),
            record.get("event_time"),
            record.get("timezone"),
            record.get("priority", "medium"),
            record.get("source_email_id"),
            "scheduled",
        ))
        return cur.lastrowid


def update_upcoming_event(event_id: int, data: dict) -> None:
    with _conn() as db:
        db.execute("""
            UPDATE upcoming_events
            SET title=?, description=?, event_date=?, event_time=?, timezone=?, priority=?
            WHERE id=?
        """, (
            data.get("title"),
            data.get("description"),
            data.get("event_date") or None,
            data.get("event_time") or None,
            data.get("timezone"),
            data.get("priority", "high"),
            event_id,
        ))


def cancel_app_events(app_id: int) -> None:
    """Mark all scheduled events for an application as cancelled."""
    with _conn() as db:
        db.execute(
            "UPDATE upcoming_events SET status='cancelled' WHERE app_id=? AND status='scheduled'",
            (app_id,),
        )


def delete_application(app_id: int) -> bool:
    """Remove an application row (e.g. a mistaken duplicate). Returns False if
    it didn't exist. Clears the FK references first (foreign_keys=ON blocks the
    delete otherwise): unlinks any email staged against it and drops its events."""
    with _conn() as db:
        db.execute("UPDATE email_staging SET matched_app_id=NULL WHERE matched_app_id=?", (app_id,))
        db.execute("DELETE FROM upcoming_events WHERE app_id=?", (app_id,))
        cur = db.execute("DELETE FROM applications WHERE id=?", (app_id,))
        return cur.rowcount > 0


def get_upcoming_events(event_type: "str | None" = None) -> list[dict]:
    with _conn() as db:
        if event_type:
            rows = db.execute("""
                SELECT ue.*, a.company, a.role
                FROM upcoming_events ue
                LEFT JOIN applications a ON a.id = ue.app_id
                WHERE ue.event_type=? AND ue.status='scheduled'
                ORDER BY ue.event_date ASC, ue.created_at ASC
            """, (event_type,)).fetchall()
        else:
            rows = db.execute("""
                SELECT ue.*, a.company, a.role
                FROM upcoming_events ue
                LEFT JOIN applications a ON a.id = ue.app_id
                WHERE ue.status='scheduled'
                ORDER BY ue.event_date ASC, ue.created_at ASC
            """).fetchall()
    return [dict(r) for r in rows]
