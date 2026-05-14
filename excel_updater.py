"""
excel_updater.py — write / update rows in job_application_tracker.xlsx.

Rules:
- Never rename or delete the file — only append/update rows.
- Dedup by Job URL before writing.
- Find the right sheet by scanning for our column headers; fall back to active sheet.
- sweep_unconfirmed() marks rows Applied > 5 days ago with no follow-up.
"""

import logging
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import load_config

log = logging.getLogger(__name__)

COLUMNS = [
    "Date Applied",
    "Company",
    "Role",
    "Location",
    "Source",
    "Match Score",
    "Interview Chance",
    "German Required",
    "Job URL",
    "Archive Folder",
    "Status",
]

STATUS_APPLIED       = "Applied ✓"
STATUS_UNCONFIRMED   = "Applied — unconfirmed (no email after 5 days)"
_UNCONFIRMED_DAYS    = 5

# Status values that count as "confirmed" — skip sweep for these
_CONFIRMED_STATUSES = {
    "Applied ✓",
    "Call Scheduled ✓",
    "Technical Scheduled ✓",
    "Final Scheduled ✓",
    "Offer Received 🎉",
    "Rejected",
}


def _open_or_create(tracker_path: Path) -> openpyxl.Workbook:
    if tracker_path.exists():
        return openpyxl.load_workbook(tracker_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return wb


def _find_sheet(wb: openpyxl.Workbook):
    """
    Return the worksheet that contains our column headers.
    Falls back to the active sheet if none match.
    """
    required = {"Company", "Job URL", "Status"}
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        header_values = {cell.value for cell in ws[1] if cell.value}
        if required.issubset(header_values):
            return ws
    return wb.active


def _ensure_headers(ws) -> dict[str, int]:
    """Return {col_name: col_index} for the header row, adding missing cols."""
    header_row = [cell.value for cell in ws[1]]
    col_map: dict[str, int] = {}
    for col_name in COLUMNS:
        if col_name in header_row:
            col_map[col_name] = header_row.index(col_name) + 1
        else:
            new_col = len(header_row) + 1
            ws.cell(row=1, column=new_col, value=col_name).font = Font(bold=True)
            header_row.append(col_name)
            col_map[col_name] = new_col
    return col_map


def _cell_value(cell) -> str:
    """Read cell value, checking hyperlink.target first."""
    if cell.hyperlink:
        return cell.hyperlink.target or ""
    return str(cell.value) if cell.value is not None else ""


def _find_row_by_url(ws, url_col: int, url: str) -> int | None:
    url = url.strip()
    for row in range(2, ws.max_row + 1):
        if _cell_value(ws.cell(row=row, column=url_col)).strip() == url:
            return row
    return None


def _write_row(ws, row: int, col_map: dict[str, int], job: dict,
               archive_path: str, status: str) -> None:
    def set_cell(col_name: str, value):
        if col_name in col_map:
            ws.cell(row=row, column=col_map[col_name], value=value)

    set_cell("Date Applied",     date.today().isoformat())
    set_cell("Company",          job.get("company", ""))
    set_cell("Role",             job.get("title", ""))
    set_cell("Location",         job.get("location", ""))
    set_cell("Source",           job.get("source", ""))
    set_cell("Match Score",      f"{job.get('match_score', '')}%")
    set_cell("Interview Chance", job.get("interview_chance", ""))
    set_cell("German Required",  job.get("german_level_required", "none"))
    set_cell("Status",           status)

    if "Job URL" in col_map:
        url = job.get("url", "")
        cell = ws.cell(row=row, column=col_map["Job URL"], value=url)
        if url:
            cell.hyperlink = url
            cell.font = Font(color="0000FF", underline="single")

    if "Archive Folder" in col_map and archive_path:
        cell = ws.cell(row=row, column=col_map["Archive Folder"], value=archive_path)
        cell.hyperlink = "file:///" + archive_path.replace("\\", "/")
        cell.font = Font(color="0000FF", underline="single")


def update_status(job: dict, status: str, cfg: dict | None = None) -> None:
    """Update only the Status cell for an existing row (matched by URL)."""
    if cfg is None:
        cfg = load_config()
    tracker_path: Path = cfg["paths"]["tracker_file"]
    if not tracker_path.exists():
        log.warning("Tracker not found — cannot update status: %s", tracker_path)
        return

    wb = openpyxl.load_workbook(tracker_path)
    ws = _find_sheet(wb)
    col_map = _ensure_headers(ws)
    url = job.get("url", "")
    row = _find_row_by_url(ws, col_map.get("Job URL", 9), url)
    if row:
        ws.cell(row=row, column=col_map["Status"], value=status)
        wb.save(tracker_path)
        log.info("Status → '%s' for %s @ %s", status, job.get("title"), job.get("company"))
    else:
        log.warning("Row not found in tracker for URL: %s", url)


def add_or_update(job: dict, archive_path: str | Path,
                  status: str = STATUS_APPLIED, cfg: dict | None = None) -> None:
    """Append a new row or update an existing one (matched by URL)."""
    if cfg is None:
        cfg = load_config()

    tracker_path: Path = cfg["paths"]["tracker_file"]
    wb = _open_or_create(tracker_path)
    ws = _find_sheet(wb)
    col_map = _ensure_headers(ws)

    url = job.get("url", "")
    existing_row = _find_row_by_url(ws, col_map.get("Job URL", 9), url) if url else None
    target_row = existing_row if existing_row else (ws.max_row + 1)

    _write_row(ws, target_row, col_map, job, str(archive_path), status)
    wb.save(tracker_path)
    action = "Updated" if existing_row else "Added"
    log.info("%s tracker row %d: %s @ %s", action, target_row, job.get("title"), job.get("company"))


def sweep_unconfirmed(cfg: dict | None = None) -> None:
    """
    Mark any 'Applied ✓' row older than 5 days as unconfirmed.
    Called once at the start of each daily run.
    """
    if cfg is None:
        cfg = load_config()

    tracker_path: Path = cfg["paths"]["tracker_file"]
    if not tracker_path.exists():
        return

    wb = openpyxl.load_workbook(tracker_path)
    ws = _find_sheet(wb)
    col_map = _ensure_headers(ws)

    date_col   = col_map.get("Date Applied")
    status_col = col_map.get("Status")
    if not date_col or not status_col:
        return

    cutoff = date.today() - timedelta(days=_UNCONFIRMED_DAYS)
    updated = 0

    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=status_col)
        if status_cell.value != STATUS_APPLIED:
            continue

        date_cell = ws.cell(row=row, column=date_col)
        raw_date  = date_cell.value
        if raw_date is None:
            continue

        try:
            if isinstance(raw_date, datetime):
                applied_date = raw_date.date()
            elif isinstance(raw_date, date):
                applied_date = raw_date
            else:
                applied_date = date.fromisoformat(str(raw_date))
        except (ValueError, TypeError):
            continue

        if applied_date <= cutoff:
            status_cell.value = STATUS_UNCONFIRMED
            updated += 1

    if updated:
        wb.save(tracker_path)
        log.info("sweep_unconfirmed: marked %d row(s) as unconfirmed", updated)


if __name__ == "__main__":
    from config import setup_logging
    cfg = load_config()
    setup_logging(cfg)
    sample = {
        "title": "Data Engineer",
        "company": "TestCorp GmbH",
        "location": "Berlin",
        "url": "https://example.com/job/42",
        "source": "LinkedIn",
        "match_score": 88,
        "interview_chance": "high",
        "german_level_required": "B1",
    }
    add_or_update(sample, r"C:\Users\f_beh\Dropbox\CV\application_history\TestCorp_DataEngineer_2026-05-14", cfg=cfg)
    sweep_unconfirmed(cfg)
    print("Test row written and sweep run.")
